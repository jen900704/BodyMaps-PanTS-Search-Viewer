from flask import Blueprint, send_file, make_response, request, jsonify, Response
from services.nifti_processor import NiftiProcessor
from services.session_manager import SessionManager, generate_uuid
from services.auto_segmentor import run_auto_segmentation
from models.application_session import ApplicationSession
from models.combined_labels import CombinedLabels
from models.base import db
from constants import Constants
import zipfile
import pandas as pd

from pathlib import Path
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm

from sqlalchemy.orm import aliased
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
import nibabel as nib
import uuid

from datetime import datetime, timedelta
from .utils import *
import requests  # ⭐ 只在這裡 import 一次 requests

# 建立 blueprint
api_blueprint = Blueprint("api", __name__)
last_session_check = datetime.now()

progress_tracker = {}  # {session_id: (start_time, expected_total_seconds)}


@api_blueprint.route("/proxy-image")
def proxy_image():
    """
    Proxy image requests so the browser only talks to our own origin.
    Front-end will call: /api/proxy-image?url=<encoded_hf_url>
    """
    raw_url = request.args.get("url")
    if not raw_url:
        return Response("Missing url parameter", status=400)

    # 可選安全限制：只允許 HuggingFace 來源
    if not raw_url.startswith("https://huggingface.co/"):
        return Response("Forbidden", status=403)

    try:
        r = requests.get(raw_url, timeout=10)
    except Exception as e:
        return Response(f"Upstream error: {e}", status=502)

    if not r.ok:
        return Response(f"Upstream status {r.status_code}", status=r.status_code)

    content_type = r.headers.get("Content-Type", "image/jpeg")

    resp = Response(r.content, status=200, mimetype=content_type)

    # ⭐ 避免 COEP 再擋圖片
    resp.headers["Cross-Origin-Resource-Policy"] = "cross-origin"

    return resp



from flask import request, jsonify
import numpy as np
import nibabel as nib
from scipy.ndimage import distance_transform_edt, label
from collections import defaultdict
from constants import Constants
import os
from openpyxl import load_workbook


SESSIONS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "tmp")
PDF_DIR = f"{Constants.PANTS_PATH}/data/pdf"
os.makedirs(SESSIONS_DIR, exist_ok=True)
os.makedirs(PDF_DIR, exist_ok=True)

def _arg(name: str, default=None):
    return request.args.get(name, default)

@api_blueprint.route('/get_preview/<clabel_ids>', methods=['GET'])
def get_preview(clabel_ids):
    # get age and thumbnail
    clabel_ids = clabel_ids.split(",")
    wb = load_workbook(os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx"))
    sheet = wb["PanTS_metadata"]
    res = {
        x: {
            "sex": "",
            "age": ""
        } for x in clabel_ids
    }
    for clabel_id in clabel_ids:
        for row in sheet.iter_rows(values_only=True):
            if row[0] == get_panTS_id(clabel_id):
                res[clabel_id]["sex"] = row[4]
                res[clabel_id]["age"] = row[5]
                break

    return jsonify(res)

# if not preloaded
@api_blueprint.route('/get_image_preview/<clabel_id>', methods=['GET'])
def get_image_preview(clabel_id):
    # get age and thumbnail
    # subfolder = "LabelTr" if int(clabel_id) < 9000 else "LabelTe"
    subfolder = "ProfileTr" if int(clabel_id) < 9000 else "ProfileTe"
    # path = os.path.join(Constants.PANTS_PATH, "data", subfolder, get_panTS_id(clabel_id), Constants.COMBINED_LABELS_FILENAME)
    # if not os.path.exists(path):
    #     print(f"File not found: {path}. Making file")
    #     npz_processor = NpzProcessor()
    #     npz_processor.combine_labels(int(clabel_id))

    path = os.path.join(Constants.PANTS_PATH, subfolder, get_panTS_id(clabel_id), "profile.jpg")
    # arr = np.load(path)["data"]
    # bytes = volume_to_png(arr)
    return send_file(
        path,
        mimetype="image/jpg",   
        as_attachment=False,
        download_name=f"{clabel_id}_slice.jpg"
    )


    

@api_blueprint.route('/get-label-colormap/<clabel_id>', methods=['GET'])
def get_label_colormap(clabel_id):
    subfolder = "LabelTr" if int(clabel_id) < 9000 else "LabelTe"
    
    clabel_path = os.path.join(Constants.PANTS_PATH, "data", subfolder, get_panTS_id(int(clabel_id)),  'combined_labels.nii.gz')

    if not os.path.exists(clabel_path):
        print(f"File not found: {clabel_path}. Making file")
        combine_label_npz(int(clabel_id))
        npzProcessor = NpzProcessor()
        npzProcessor.npz_to_nifti(int(clabel_id))
    try:
        clabel_array = nib.load(clabel_path)
        clabel_array = clabel_array.get_fdata()
        print("[DEBUG] Nifti loaded, shape =", clabel_array.shape)

        filled_array = fill_voids_with_nearest_label(clabel_array)
        print("[DEBUG] fill_voids_with_nearest_label done")

        adjacency = build_adjacency_graph(filled_array)
        print("[DEBUG] build_adjacency_graph done")

        unique_labels = sorted(adjacency.keys())
        color_map, color_usage_count = assign_colors_with_high_contrast(unique_labels, adjacency)
        print("[DEBUG] Color map generated:", color_map, color_usage_count)

        return jsonify(color_map)

    except Exception as e:
        print("[❌ EXCEPTION]", str(e))
        return jsonify({"error": str(e)}), 500




# @api_blueprint.before_request
# def before_request():
#     global last_session_check
#     current_time = datetime.now()
#     if current_time >= last_session_check + timedelta(minutes=Constants.SCHEDULED_CHECK_INTERVAL):
#         session_manager = SessionManager.instance()
#         expired = session_manager.get_expired()
#         for app_session in expired:
#             session_manager.terminate_session(app_session.session_id)
        
#         last_session_check = current_time

@api_blueprint.route('/', methods=['GET'])
def home():
    return "api"


@api_blueprint.route('/upload', methods=['POST'])
def upload():
    try:
        session_id = request.form.get('SESSION_ID')
        if not session_id:
            return jsonify({"error": "No session ID provided"}), 400
        
        base_path = os.path.join(Constants.SESSIONS_DIR_NAME, session_id)
        os.makedirs(base_path, exist_ok=True)

        nifti_multi_dict = request.files
        filenames = list(nifti_multi_dict)
        main_nifti = nifti_multi_dict.get(Constants.MAIN_NIFTI_FORM_NAME)

        if main_nifti:
            main_nifti_path = os.path.join(base_path, Constants.MAIN_NIFTI_FILENAME)
            main_nifti.save(main_nifti_path)
            filenames.remove(Constants.MAIN_NIFTI_FORM_NAME)
        else:
            return jsonify({"error": "Main NIFTI file missing"}), 400

        nifti_processor = NiftiProcessor.from_clabel_path(os.path.join(base_path, Constants.COMBINED_LABELS_FILENAME))

        combined_labels, organ_intensities = nifti_processor.combine_labels(filenames, nifti_multi_dict, save=True)

        resp = {
            'status': "200",
            'session_id': session_id,
            'organ_intensities': organ_intensities
        }
        return jsonify(resp)
    except Exception as e:
        print(f"❌ [Upload Error] {e}")
        return jsonify({"error": "Internal server error"}), 500

@api_blueprint.route('/mask-data', methods=['POST'])
def get_mask_data():
    session_key = request.form.get('sessionKey')
    if not session_key:
        return jsonify({"error": "Missing sessionKey"}), 400

    result = get_mask_data_internal(session_key)
    return jsonify(result)

  
@api_blueprint.route('/get-main-nifti/<clabel_id>', methods=['GET'])
def get_main_nifti(clabel_id):
    subfolder = "ImageTr" if int(clabel_id) < 9000 else "ImageTe" 
    main_nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(clabel_id)}/{Constants.MAIN_NIFTI_FILENAME}"

    if os.path.exists(main_nifti_path):
        response = make_response(send_file(main_nifti_path, mimetype='application/gzip'))

        response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
        response.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'
        response.headers['Content-Encoding'] = 'gzip'

    else:
        print(f"Could not find filepath: {main_nifti_path}. ")
        return jsonify({"error": "Could not find filepath"}), 404
        
        # npz_path = main_nifti_path.replace(".nii.gz", ".npz")
        # if not os.path.exists(npz_path):   
        #     return jsonify({"error": "Could not find npz filepath"}), 404
        # npz_processor = NpzProcessor()
        # npz_processor.npz_to_nifti(int(clabel_id), combined_label=False, save=True)  
        
        # response = make_response(send_file(main_nifti_path, mimetype='application/gzip'))

        # response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
        # response.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'
        # response.headers['Content-Encoding'] = 'gzip'

    return response




@api_blueprint.route('/get-report/<id>', methods=['GET'])
def get_report(id):
    temp_pdf_path = f"{PDF_DIR}/temp.pdf"
    output_pdf_path = f"{PDF_DIR}/final.pdf"
    try:
        try:
            organ_metrics = get_mask_data_internal(id)
            organ_metrics = organ_metrics.get("organ_metrics", [])
        except Exception as e:
            return jsonify({"error": f"Error loading organ metrics: {str(e)}"}), 500

        subfolder = "ImageTr" if int(id) < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if int(id) < 9000 else "LabelTe"

        base_path = f"{SESSIONS_DIR}/{id}"
        ct_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/{Constants.MAIN_NIFTI_FILENAME}"
        masks = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        
        npz_processor = NpzProcessor()

        # if (not os.path.exists(ct_path)):
        #     npz_processor.npz_to_nifti(int(id), combined_label=False, save=True)

        if (not os.path.exists(masks)): 
            npz_processor.combine_labels(int(id), keywords={"pancrea": "pancreas"}, save=True)
            npz_processor.npz_to_nifti(int(id), combined_label=True, save=True)
            
        template_pdf = os.getenv("TEMPLATE_PATH", "report_template_3.pdf")

        extracted_data = None
        column_headers = None
        try:
            csv_path = f"{base_path}/info.csv"
            df = pd.read_csv(csv_path)
            extracted_data = df.iloc[0] if len(df) > 0 else None
            column_headers = df.columns.tolist()
        except Exception:
            pass

        generate_pdf_with_template(
            output_pdf=output_pdf_path,
            folder_name=id,
            ct_path=ct_path,
            mask_path=masks,
            template_pdf=template_pdf,
            temp_pdf_path=temp_pdf_path,
            id=id,
            extracted_data=extracted_data,
            column_headers=column_headers
        )

        return send_file(
            output_pdf_path,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"report_{id}.pdf"
        )

    except Exception as e:
        return jsonify({"error": f"Unhandled error: {str(e)}"}), 500

    finally:
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)


@api_blueprint.route('/get-segmentations/<combined_labels_id>', methods=['GET'])
async def get_segmentations(combined_labels_id):
    subfolder = "LabelTr" if int(combined_labels_id) < 9000 else "LabelTe" 
    nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(combined_labels_id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
    labels = list(Constants.PREDEFINED_LABELS.values()) 
    if not os.path.exists(nifti_path):
        await store_files(combined_labels_id)
        niftiProcessor = NpzProcessor()
        niftiProcessor.nifti_combine_labels(int(combined_labels_id))
        # print(f"Could not find filepath: {nifti_path}. Creating a new one")
        # npz_path = nifti_path.replace(".nii.gz", ".npz")
        # npz_processor = NpzProcessor()
        # if not os.path.exists(npz_path):   
        #     print(f"Could not find npz filepath: {npz_path}. Creating a new one")

        #     # ! pancrea instead of pancreas to include pancreatic labels
        #     npz_processor.combine_labels(combined_labels_id, keywords={"pancrea": "pancreas"}, save=True)
            
        # npz_processor.npz_to_nifti(int(combined_labels_id), combined_label=True, save=True)   

    img = nib.load(nifti_path)
    data = img.get_fdata()
    if img.get_data_dtype() != np.uint8:
        print("⚠️ Detected float label map, converting to uint8 for Niivue compatibility...")

    try:
        img = nib.load(nifti_path)
        data = img.get_fdata()

        if img.get_data_dtype() != np.uint8:
            
            data_uint8 = data.astype(np.uint8)
            new_img = nib.Nifti1Image(data_uint8, img.affine, header=img.header)
            new_img.set_data_dtype(np.uint8)

            converted_path = nifti_path#.replace(".nii.gz", "_uint8.nii.gz")

            if not os.path.exists(converted_path):
                nib.save(new_img, converted_path)
        else:
            converted_path = nifti_path

        response = make_response(send_file(converted_path, mimetype='application/gzip'))
        response.headers['Cross-Origin-Opener-Policy'] = 'same-origin'
        response.headers['Cross-Origin-Embedder-Policy'] = 'require-corp'
        response.headers['Content-Encoding'] = 'gzip'

        return response

    except Exception as e:
        print(f"❌ [get-segmentations ERROR] {e}")
        return jsonify({"error": str(e)}), 500


@api_blueprint.route('/download/<id>', methods=['GET'])
def download_segmentation_zip(id):
    try:
        subfolder = "LabelTr" if int(id) < 9000 else "LabelTe"
        outputs_ct_folder = Path(f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/segmentations")
        
        if not os.path.exists(outputs_ct_folder):
            return jsonify({"error": "Outputs/ct folder not found"}), 404
        
        files = list(outputs_ct_folder.glob("*"))

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for file_path in files:
                zip_file.write(file_path, arcname=file_path.name) 

        zip_buffer.seek(0)  # rewind

        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name=f"case_{id}_segmentations.zip"
        )



    except Exception as e:
        print(f"❌ [Download Error] {e}")
        return jsonify({"error": "Internal server error"}), 500

import threading
import time

@api_blueprint.route('/auto_segment/<session_id>', methods=['POST'])
def auto_segment(session_id):

    if 'MAIN_NIFTI' not in request.files:
        return jsonify({"error": "No CT file provided"}), 400

    ct_file = request.files['MAIN_NIFTI']
    model_name = request.form.get("MODEL_NAME", None)

    # Check if model name is valid
    if model_name is None:
        return {"error": "MODEL_NAME is required."}, 400
    # Step 1: Create a unique session directory to store CT and mask
    session_path = os.path.join(SESSIONS_DIR, session_id)
    os.makedirs(session_path, exist_ok=True)

    # Step 2: Save CT file under this session
    input_path = os.path.join(session_path, ct_file.filename)
    ct_file.save(input_path)

    def do_segmentation_and_zip():
        time.sleep(10)
        output_mask_dir = run_auto_segmentation(input_path, session_dir=session_path, model=model_name)

        if output_mask_dir is None or not os.path.exists(output_mask_dir):
            print(f"❌ Auto segmentation failed for session {session_id}")
            return ##the logic still needs to be improved in the future. when output_mask_dir is none here, no error output at user's end

        zip_path = os.path.join(session_path, "auto_masks.zip")
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for filename in os.listdir(output_mask_dir):
                if filename.endswith(".nii.gz"):
                    abs_path = os.path.join(output_mask_dir, filename)
                    zipf.write(abs_path, arcname=filename)

        start_time, expected_time, _ = progress_tracker[session_id]
        progress_tracker[session_id] = (start_time, expected_time, True)
        progress_tracker.pop(session_id, None)

        
        
        print(f"✅ Finished segmentation and zipping for session {session_id}")

    #threading.Thread(target=do_segmentation_and_zip).start()
    threading.Thread(target=do_segmentation_and_zip, ).start()
    print("[Server] auto_segment request is returning now")
    return jsonify({"message": "Segmentation started"}), 200



@api_blueprint.route('/get_result/<session_id>', methods=['GET'])
def get_result(session_id):
    session_path = os.path.join(SESSIONS_DIR, session_id)
    zip_path = os.path.join(session_path, "auto_masks.zip")

    wait_for_file(zip_path, timeout=30)

    response = send_file(
        zip_path,
        as_attachment=True,
        download_name="auto_masks.zip"
    )
    response.headers["X-Session-Id"] = session_id
    return response

#
#@api_blueprint.route('/progress_end/<session_id>', methods=['GET'])
#def progress_end(session_id):
#    progress_tracker.pop(session_id, None)
#    return jsonify({"message": "Progress End"}), 200

@api_blueprint.route('/ping', methods=['GET'])
def ping():
    return jsonify({"message": "pong"}), 200

@api_blueprint.route("/search", methods=["GET"])
def api_search():
    # return jsonify({"message": "pong"}), 200
    df = apply_filters(DF).copy()
    sort_by  = (_arg("sort_by", "top") or "top").strip().lower()
    sort_by  = (_arg("sort_by", "top") or "top").strip().lower()
    df = ensure_sort_cols(df)

    # ---- 排序參數 ----
    sort_by  = (_arg("sort_by", "top") or "top").strip().lower()
    sort_dir = (_arg("sort_dir", "asc") or "asc").strip().lower()

    if sort_by in ("top", "quality"):
        by  = ["__complete", "__spacing_sum", "__shape_sum", "__case_sortkey"]
        asc = [False, True, False, True]
    elif sort_by in ("id", "id_asc"):
        by, asc = ["__case_sortkey"], [True]
    elif sort_by == "id_desc":
        by, asc = ["__case_sortkey"], [False]
    elif sort_by in ("shape_desc", "shape"):
        by, asc = ["__shape_sum", "__case_sortkey"], [False, True]
    elif sort_by in ("spacing_asc", "spacing"):
        by, asc = ["__spacing_sum", "__case_sortkey"], [True, True]
    elif sort_by == "age_asc":
        by, asc = ["__age", "__case_sortkey"], [True, True]
    elif sort_by == "age_desc":
        by, asc = ["__age", "__case_sortkey"], [False, True]
    else:
        key_map = {"id": "__case_sortkey", "spacing": "__spacing_sum", "shape": "__shape_sum"}
        k = key_map.get(sort_by, "__case_sortkey")
        by, asc = [k, "__case_sortkey"], [(sort_dir != "desc"), True]

    # ---- 排序 ----
    df = df.sort_values(by=by, ascending=asc, na_position="last", kind="mergesort")

    # ---- 分頁：注意 total 先算完篩選後的完整筆數 ----
    total    = int(len(df))
    page     = max(to_int(_arg("page", "1")) or 1, 1)
    per_page = to_int(_arg("per_page", "24")) or 24
    per_page = max(1, min(per_page, 1_000_000))

    pages = max(1, int(math.ceil(total / per_page)))
    page  = max(1, min(page, pages))
    start, end = (page - 1) * per_page, (page - 1) * per_page + per_page

    # ---- 轉成前端想要的 items ----
    items = [row_to_item(r) for _, r in df.iloc[start:end].iterrows()]
    items = clean_json_list(items)

    return jsonify({
        "items": items,         # ← 前端只讀這個渲染卡片
        "total": total,         # ← 正確的最終數量
        "page": page,
        "per_page": per_page,
        "query": request.query_string.decode(errors="ignore") or ""
    })


def _facet_counts_with_unknown(df: pd.DataFrame, col_key: str, top_k: int = 6) -> Dict[str, Any]:
    """Compute facet rows + unknown count, with robust handling for NaN/strings."""
    rows: List[Dict[str, Any]] = []
    unknown: int = 0

    key_to_col = {
        "ct_phase": ("__ct", str),
        "manufacturer": ("__mfr", str),
        "year": ("__year_int", int),
        "sex": ("__sex", str),
        "tumor": ("__tumor01", int),
        "model": ("model", str),
        "study_type": ("study_type", str),
        "site_nat": ("site_nationality", str),
        "site_nationality": ("site_nationality", str),
    }
    if col_key not in key_to_col:
        return {"rows": [], "unknown": 0}

    col_name, _typ = key_to_col[col_key]
    if col_name not in df.columns:
        return {"rows": [], "unknown": 0}

    ser = df[col_name]

    # ---- Year：數值化、NaN 視為 unknown ----
    if col_key == "year":
        s_num = pd.to_numeric(ser, errors="coerce")
        unknown = int(s_num.isna().sum())
        vc = s_num.dropna().astype(int).value_counts()
        rows = [{"value": int(v), "count": int(c)} for v, c in vc.items()]
        rows.sort(key=lambda x: (-x["count"], x["value"]))
        if top_k and top_k > 0:
            rows = rows[:top_k]
        return {"rows": rows, "unknown": unknown}

    # ---- 其他欄位：把空字串/unknown 類型歸入 unknown ----
    s_str = ser.astype(str).str.strip()
    s_lc = s_str.str.lower()
    unknown_mask = ser.isna() | (s_str == "") | (s_lc.isin({"unknown", "nan", "none", "n/a", "na"}))
    unknown = int(unknown_mask.sum())

    vals = ser[~unknown_mask]
    vc = vals.value_counts(dropna=False)

    tmp_rows: List[Dict[str, Any]] = []
    for v, c in vc.items():
        if col_key == "tumor":
            # tumor 僅接受 0/1
            try:
                iv = int(v)
            except Exception:
                continue
            if iv not in (0, 1):
                continue
            tmp_rows.append({"value": iv, "count": int(c)})
        else:
            tmp_rows.append({"value": v, "count": int(c)})

    # 排序：count desc，再 value 升（字串比較避免型別問題）
    tmp_rows.sort(key=lambda x: (-x["count"], str(x["value"])))
    if top_k and top_k > 0:
        tmp_rows = tmp_rows[:top_k]

    rows = tmp_rows
    return {"rows": rows, "unknown": unknown}


def _prune_zero_rows(rows: List[Dict[str, Any]], keep_zero: bool) -> List[Dict[str, Any]]:
    """依需求濾掉 count<=0；當 keep_zero=True（對應 guarantee=1）則不濾。"""
    if keep_zero:
        return rows
    out: List[Dict[str, Any]] = []
    for r in rows or []:
        try:
            c = int(r.get("count") or 0)
        except Exception:
            c = 0
        if c > 0:
            out.append(r)
    return out


@api_blueprint.route("/facets", methods=["GET"])
def api_facets():
    try:
        fields_raw = (_arg("fields","ct_phase,manufacturer") or "").strip()
        fields = [f.strip().lower() for f in fields_raw.split(",") if f.strip()]

        valid  = {
            "ct_phase","manufacturer","year","sex","tumor",
            "model","study_type","site_nat","site_nationality"
        }
        fields = [f for f in fields if f in valid] or ["ct_phase","manufacturer"]
        top_k  = to_int(_arg("top_k","6")) or 6
        guarantee = (_arg("guarantee","0") or "0").strip().lower() in ("1","true","yes","y")

        # 先應用目前的過濾條件
        df_now = apply_filters(DF)
        base_for_ranges = df_now if len(df_now) else DF

        facets: Dict[str, List[Dict[str, Any]]] = {}
        unknown_counts: Dict[str, int] = {}

        # 為每個 facet 準備自我排除的條件（避免自我影響）
        exclude_map = {
            "ct_phase": {"ct_phase"},
            "manufacturer": {"manufacturer","mfr_is_null","manufacturer_is_null"},
            "year": {"year_from","year_to"},
            "sex": {"sex"},
            "tumor": {"tumor"},
            "model": {"model"},
            "study_type": {"study_type"},
            "site_nat": {"site_nat","site_nationality"},
            "site_nationality": {"site_nat","site_nationality"},
        }

        for f in fields:
            ex = exclude_map.get(f, set())
            # 若 guarantee=1 且目前篩完為空，改用全量 DF 以「保證列出所有可能值」
            src = (DF if (guarantee and len(df_now) == 0) else df_now)
            df_facet = apply_filters(src, exclude=ex)
            res = _facet_counts_with_unknown(df_facet, f, top_k=top_k)

            # guarantee=0 時砍掉 count<=0 的項目
            rows = _prune_zero_rows(res.get("rows") or [], keep_zero=guarantee)
            facets[f] = rows
            unknown_counts[f] = int(res.get("unknown") or 0)

        # 年齡/年份範圍（原樣保留）
        def _minmax(series: pd.Series):
            s = series.dropna()
            if not len(s): return (None, None)
            return (float(s.min()), float(s.max()))

        age_min = age_max = None
        year_min = year_max = None
        if "__age" in base_for_ranges:
            age_min, age_max = _minmax(base_for_ranges["__age"])
        if "__year_int" in base_for_ranges:
            yr = base_for_ranges["__year_int"].dropna().astype(int)
            if len(yr):
                year_min, year_max = int(yr.min()), int(yr.max())

        return jsonify({
            "facets": facets,
            "unknown_counts": unknown_counts,
            "age_range": {"min": age_min, "max": age_max},
            "year_range": {"min": year_min, "max": year_max},
            "total": int(len(df_now)),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
@api_blueprint.route("/random", methods=['GET'])
def api_random_topk_rotate_norand():
    """
    推薦：完整資料優先 → 取 Top-K(預設100) → 環狀位移 → 可排除最近看過
    排序：__spacing_sum ↑, __shape_sum ↓, __case_sortkey ↑
    """
    try:
        scope = (request.args.get("scope", "filtered") or "filtered").strip().lower()
        base_df = apply_filters(DF)
        if len(base_df) == 0 and scope == "all":
            base_df = DF.copy()

        base_df = ensure_sort_cols(base_df)

        # 只取完整資料；若沒有完整的就退回全部
        df_full = base_df[base_df["__complete"]] if "__complete" in base_df.columns else base_df
        if len(df_full) == 0:
            df_full = base_df
        df = df_full.sort_values(
            by=["__spacing_sum","__shape_sum","__case_sortkey"],
            ascending=[True, False, True],
            na_position="last",
            kind="mergesort",
        )

        if len(df) == 0:
            return jsonify({"items": [], "total": 0, "meta": {"k": 0, "used_recent": 0}}), 200

        # n, k
        try: n = int(request.args.get("n") or 3)
        except Exception: n = 3
        n = max(1, min(n, len(df)))

        try: K = int(request.args.get("k") or 100)
        except Exception: K = 100
        K = max(n, min(K, len(df)))

        # recent 排除
        recent_raw = (request.args.get("recent") or "").strip()
        used_recent = 0
        if recent_raw:
            recent_ids = {s.strip() for s in recent_raw.split(",") if s.strip()}
            key = df["__case_str"].astype(str) if "__case_str" in df.columns else None
            if key is not None:
                mask = ~key.isin(recent_ids)
                used_recent = int((~mask).sum())
                df2 = df[mask]
                if len(df2): df = df2

        topk = df.iloc[:K]
        if len(topk) == 0:
            return jsonify({"items": [], "total": 0, "meta": {"k": 0, "used_recent": used_recent}}), 200

        off_arg = request.args.get("offset")
        if off_arg is not None:
            try: offset = int(off_arg) % len(topk)
            except Exception: offset = 0
        else:
            now = datetime.utcnow()
            offset = ((now.minute * 60) + now.second) % len(topk)

        idx = list(range(len(topk))) + list(range(len(topk)))
        pick = idx[offset:offset + min(n, len(topk))]
        sub = topk.iloc[pick]

        items = [row_to_item(r) for _, r in sub.iterrows()]
        resp = jsonify({
            "items": clean_json_list(items),
            "total": int(len(df)),
            "meta": {"k": int(len(topk)), "used_recent": used_recent, "offset": int(offset)}
        })
        r = make_response(resp)
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        r.headers["Pragma"] = "no-cache"
        r.headers["Expires"] = "0"
        return r

    except Exception as e:
        return jsonify({"error": str(e)}), 400

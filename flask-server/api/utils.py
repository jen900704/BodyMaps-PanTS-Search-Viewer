from flask import Blueprint, send_file, make_response, request, jsonify
from services.nifti_processor import NiftiProcessor
from services.session_manager import SessionManager, generate_uuid
from services.auto_segmentor import run_auto_segmentation
from models.application_session import ApplicationSession
from models.combined_labels import CombinedLabels
from models.base import db
from constants import Constants

from io import BytesIO
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

from typing import Any, Dict, Optional, Set, List, Tuple

import os
import uuid
import re
import time
import math
import numpy as np
import nibabel as nib
from scipy.ndimage import distance_transform_edt
from collections import defaultdict
from services.npz_processor import NpzProcessor
from PIL import Image
from openpyxl import load_workbook
import requests
import pandas as pd
# Track last session validation time
last_session_check = datetime.now()

# Progress tracking structure: {session_id: (start_time, expected_total_seconds)}
progress_tracker = {}

def id_is_training(index):
    return index < 9000



def combine_label_npz(index: int):
    npz_processor = NpzProcessor()
    npz_processor.combine_labels(index)
    return
def get_panTS_id(index):
    cur_case_id = str(index)
    iter = max(0, 8 - len(str(index)))
    for _ in range(iter):
        cur_case_id = "0" + cur_case_id
    cur_case_id = "PanTS_" + cur_case_id    
    return cur_case_id

def clean_nan(obj):
    """Recursively replace NaN with None for JSON serialization."""
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan(elem) for elem in obj]
    elif isinstance(obj, float) and math.isnan(obj):
        return None
    else:
        return obj

def format_value(value):
    """Format values for display, replacing 999999 or None with 'N/A'."""
    return "N/A" if value in [999999, None] else str(value)

def organname_to_name(filename):
    """Convert a NIfTI file name to a human-readable organ name."""
    name = filename.replace(".nii.gz", "").replace("_", " ")
    return name.title()

def get_mask_data_internal(id, fallback=False):
    """Retrieve or compute organ metadata from NIfTI and mask paths for a session."""
    try:
        subfolder = "ImageTr" if int(id) < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if int(id) < 9000 else "LabelTe"
        main_nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(id)}/{Constants.MAIN_NIFTI_FILENAME}"
        combined_labels_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        print(f"[INFO] Processing NIFTI for id {id}")
        organ_intensities = None
        
        organ_intensities_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(id)}/{Constants.ORGAN_INTENSITIES_FILENAME}"
        if not os.path.exists(organ_intensities_path) or not os.path.exists(combined_labels_path):
            npz_processor = NpzProcessor()
            labels, organ_intensities = npz_processor.combine_labels(int(id), keywords={"pancrea": "pancreas"}, save=True)
        else: 
            with open(organ_intensities_path, "r") as f:
                organ_intensities = json.load(f)
        
        nifti_processor = NiftiProcessor(main_nifti_path, combined_labels_path)
        nifti_processor.set_organ_intensities(organ_intensities)
        organ_metadata = nifti_processor.calculate_metrics()
        organ_metadata = clean_nan(organ_metadata)

        return organ_metadata

    except Exception as e:
        print(f"[ERROR] get_mask_data_internal: {e}")
        return {"error": str(e)}

def generate_distinct_colors(n):
    """Generate n visually distinct RGB colors."""
    import colorsys
    HSV_tuples = [(x / n, 0.7, 0.9) for x in range(n)]
    RGB_tuples = [tuple(int(c * 255) for c in colorsys.hsv_to_rgb(*hsv)) for hsv in HSV_tuples]
    return RGB_tuples

def fill_voids_with_nearest_label(label_array):
    """Fill all 0-valued voxels with the nearest non-zero label."""
    mask = label_array == 0
    if not np.any(mask):
        return label_array

    nonzero_coords = np.array(np.nonzero(label_array)).T
    distances, indices = distance_transform_edt(mask, return_indices=True)
    filled_array = label_array.copy()
    filled_array[mask] = label_array[tuple(indices[:, mask])]
    return filled_array

def build_adjacency_graph(label_array):
    """Build adjacency graph of label connectivity in 6 directions."""
    adjacency = defaultdict(set)
    offsets = [(-1, 0, 0), (1, 0, 0),
               (0, -1, 0), (0, 1, 0),
               (0, 0, -1), (0, 0, 1)]

    for dx, dy, dz in offsets:
        shifted = np.roll(label_array, shift=(dx, dy, dz), axis=(0, 1, 2))
        mask = (label_array != shifted) & (label_array != 0) & (shifted != 0)
        l1 = label_array[mask]
        l2 = shifted[mask]
        for a, b in zip(l1, l2):
            if a != b:
                adjacency[a].add(b)
                adjacency[b].add(a)
    return adjacency

def assign_colors_with_high_contrast(label_ids, adjacency_graph, min_initial_colors=20, max_total_colors=50):
    """
    Assign colors to labels such that adjacent labels have different colors,
    maximizing contrast and balance.
    """
    from itertools import combinations
    import colorsys

    def generate_distinct_colors(n):
        HSV_tuples = [(x / n, 0.7, 0.9) for x in range(n)]
        RGB_tuples = [tuple(int(c * 255) for c in colorsys.hsv_to_rgb(*hsv)) for hsv in HSV_tuples]
        return RGB_tuples

    def can_use_color(label, color_idx, assignments, adjacency_graph):
        for neighbor in adjacency_graph[label]:
            if assignments.get(neighbor) == color_idx:
                return False
        return True

    label_ids = sorted(label_ids)
    assignments = {}
    num_colors = min_initial_colors
    color_usage_count = {i: 0 for i in range(num_colors)}

    while True:
        colors = generate_distinct_colors(num_colors)
        assignments.clear()
        color_usage_count = {i: 0 for i in range(num_colors)}
        success = True

        for label in label_ids:
            color_order = sorted(range(num_colors), key=lambda c: (color_usage_count[c], c))
            for color_idx in color_order:
                if can_use_color(label, color_idx, assignments, adjacency_graph):
                    assignments[label] = color_idx
                    color_usage_count[color_idx] += 1
                    break
            else:
                success = False
                break

        if success:
            break
        elif num_colors >= max_total_colors:
            print(f"⚠️ Warning: reached max color count {max_total_colors}, some neighbors may share color")
            break
        else:
            num_colors += 1

    final_colors = generate_distinct_colors(num_colors)
    print(f"✅ Final color count used: {len(set(assignments.values()))}")

    color_map = {
        str(round(label)): {
            "R": final_colors[color_idx][0],
            "G": final_colors[color_idx][1],
            "B": final_colors[color_idx][2],
            "A": 128
        }
        for label, color_idx in assignments.items()
    }

    return color_map, color_usage_count

def wait_for_file(filepath, timeout=30, check_interval=0.5):
    """Wait until a file exists, or timeout is reached."""
    start_time = time.time()
    while not os.path.exists(filepath):
        if time.time() - start_time > timeout:
            raise TimeoutError(f"Timeout: File {filepath} not found after {timeout} seconds.")
        time.sleep(check_interval)

def volume_to_png(volume, axis=2, index=None):
    if index is None:
        index = volume.shape[axis] // 2
    
    slice_ = np.take(volume, index, axis=axis)
    # window_center = 40 
    # window_width = 400 
    # min_val = window_center - window_width / 2
    # max_val = window_center + window_width / 2

    # slice_clipped = np.clip(slice_, min_val, max_val)
    # slice_norm = 255 * (slice_clipped - min_val) / (max_val - min_val)
    slice_norm = 255 * (slice_ - np.min(slice_)) / (np.max(slice_) - np.min(slice_))
    slice_norm = slice_norm.astype(np.uint8)
    
    slice_norm = np.rot90(slice_norm, k=1)
    slice_norm = np.flip(slice_norm, axis=0)

    pil_img = Image.fromarray(slice_norm)
    buf = BytesIO()
    pil_img.save(buf, format="PNG")
    buf.seek(0)
    return buf
def generate_pdf_with_template(
    output_pdf,
    folder_name,
    ct_path,
    mask_path,
    template_pdf,
    temp_pdf_path,
    id,
    extracted_data=None,
    column_headers=None,
):
    import os
    import nibabel as nib
    import numpy as np
    import pandas as pd
    from PyPDF2 import PdfReader, PdfWriter
    from PyPDF2._page import PageObject
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    LABELS = {v: k for k, v in Constants.PREDEFINED_LABELS.items()}
    NAME_TO_ORGAN = {
        # Pancreas and its lesions
        "pancreas": "pancreas",
        "pancreas_body": "pancreas",
        "pancreas_head": "pancreas",
        "pancreas_tail": "pancreas",
        "pancreatic_lesion": "pancreas",
        "pancreatic_duct": "pancreas",

        # All other organs: map to self
        "aorta": "aorta",
        "adrenal_gland_left": "adrenal_gland_left",
        "adrenal_gland_right": "adrenal_gland_right",
        "bladder": "bladder",
        "common_bile_duct": "common_bile_duct",
        "celic_artery": "celiac_artery",
        "colon": "colon",
        "duodenum": "duodenum",
        "femur_right": "femur_right",
        "femur_left": "femur_left",
        "gall_bladder": "gall_bladder",
        "postcava": "postcava",
        "kidney_left": "kidney_left",
        "kidney_right": "kidney_right",
        "liver": "liver",
        "postcava": "postcava",
        "prostate": "prostate",
        "superior_mesenteric_artery": "superior_mesenteric_artery",
        "intestine": "intestine",
        "spleen": "spleen",
        "stomach": "stomach",
        "veins": "veins",
    }

    try:
        temp_pdf = canvas.Canvas(temp_pdf_path, pagesize=letter)
        width, height = letter
        left_margin, top_margin = 50, 100
        line_height, section_spacing = 12, 30
        y_position = height - top_margin

        def reset_page():
            nonlocal y_position
            temp_pdf.showPage()
            y_position = height - 120
            temp_pdf.setFont("Helvetica", 10)

        def write_wrapped_text(x, y, content, bold=False, font_size=10, max_width=None):
            temp_pdf.setFont("Helvetica-Bold" if bold else "Helvetica", font_size)
            words = content.split()
            current_line = ""
            max_width = max_width or width - left_margin * 2
            for word in words:
                if temp_pdf.stringWidth(current_line + word + " ", "Helvetica", font_size) > max_width:
                    temp_pdf.drawString(x, y, current_line.strip())
                    y -= line_height
                    current_line = f"{word} "
                    if y < 50:
                        reset_page()
                        y = y_position
                else:
                    current_line += f"{word} "
            if current_line:
                temp_pdf.drawString(x, y, current_line.strip())
                y -= line_height
            return y

        def safe_extract(index, default="N/A"):
            if extracted_data is not None and index in extracted_data:
                val = extracted_data[index]
                return "N/A" if pd.isna(val) else val
            return default
        
        wb = load_workbook(os.path.join(Constants.PANTS_PATH, "data", "metadata.xlsx"))
        sheet = wb["PanTS_metadata"]
        age = None
        sex = "-"
        contrast = ""
        study_detail = ""
        for row in sheet.iter_rows(values_only=True):
            if row[0] == get_panTS_id(folder_name):
                age = row[5]
                sex = row[4]
                contrast = row[3]
                study_detail = row[8]
                break

        # Title
        temp_pdf.setFont("Helvetica-Bold", 26)
        title_text = "MEDICAL REPORT"
        title_width = temp_pdf.stringWidth(title_text, "Helvetica-Bold", 26)
        temp_pdf.drawString((width - title_width) / 2, height - 70, title_text)
        y_position = height - 100

        # Patient info
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "PATIENT INFORMATION")
        y_position -= line_height

        left_y = write_wrapped_text(left_margin, y_position, f"PANTS ID: {folder_name}")
        right_y = write_wrapped_text(width / 2, y_position, f"Sex: {sex}")
        y_position -= line_height
        
        write_wrapped_text(left_margin, y_position, f"Age: {age}")
        
        y_position = min(left_y, right_y) - section_spacing

        # Imaging detail
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "IMAGING DETAIL")
        y_position -= line_height

        ct_nii = nib.load(ct_path)
        spacing = ct_nii.header.get_zooms()
        shape = ct_nii.shape

        try:
            scanner_info = str(ct_nii.header['descrip'].tobytes().decode('utf-8')).strip().replace('\x00', '')
        except Exception:
            scanner_info = "N/A"


        y_position = write_wrapped_text(left_margin, y_position, f"Spacing: {spacing}")
        y_position = write_wrapped_text(left_margin, y_position, f"Shape: {shape}")
        y_position = write_wrapped_text(left_margin, y_position, f"Study type: {study_detail}")
        y_position = write_wrapped_text(left_margin, y_position, f"Contrast: {contrast}")
        y_position -= section_spacing

        # Load image data
        ct_array = ct_nii.get_fdata()
        mask_array = nib.load(mask_path).get_fdata().astype(np.uint8)
        voxel_volume = np.prod(nib.load(mask_path).header.get_zooms()) / 1000  # mm³ to cm³
        print(np.unique(mask_array))

        # AI Measurements
        temp_pdf.setFont("Helvetica-Bold", 12)
        temp_pdf.drawString(left_margin, y_position, "AI MEASUREMENTS")
        y_position -= line_height

        # Table configuration
        headers = ["Organ", "Volume (cc)", "Mean HU"]
        col_widths = [120, 100, 100]
        row_height = 20

        def draw_table_row(row_data, is_header=False):
            nonlocal y_position
            if y_position - row_height < 50:
                reset_page()
                temp_pdf.setFont("Helvetica-Bold", 12)
                temp_pdf.drawString(left_margin, y_position, "AI MEASUREMENTS (continued)")
                y_position -= line_height
                draw_table_row(headers, is_header=True)
            x = left_margin
            temp_pdf.setFont("Helvetica-Bold" if is_header else "Helvetica", 9)
            for i, cell in enumerate(row_data):
                temp_pdf.drawString(x + 2, y_position - row_height + 5, str(cell))
                temp_pdf.line(x, y_position, x, y_position - row_height)
                x += col_widths[i]
            temp_pdf.line(left_margin + sum(col_widths), y_position, left_margin + sum(col_widths), y_position - row_height)
            temp_pdf.line(left_margin, y_position, left_margin + sum(col_widths), y_position)
            y_position -= row_height
            temp_pdf.line(left_margin, y_position, left_margin + sum(col_widths), y_position)

        draw_table_row(headers, is_header=True)

        lession_volume_dict={}
        for organ, label_id in LABELS.items():
            if organ in NAME_TO_ORGAN and NAME_TO_ORGAN[organ] != organ:
                mask = (mask_array == label_id)
                if not np.any(mask):
                    print("none")
                    continue
                volume = np.sum(mask) * voxel_volume
                mean_hu = np.mean(ct_array[mask])
                if NAME_TO_ORGAN[organ] in lession_volume_dict:
                    lession_volume_dict[NAME_TO_ORGAN[organ]]["number"] += 1
                    lession_volume_dict[NAME_TO_ORGAN[organ]]["volume"] += volume
                else:
                    lession_volume_dict[NAME_TO_ORGAN[organ]] = {
                        "number": 1,
                        "volume": volume
                    }
                    
        print(lession_volume_dict)
        
        for organ, label_id in LABELS.items():
            if organ in NAME_TO_ORGAN and NAME_TO_ORGAN[organ] != organ:
                continue
            if label_id == 0:
                continue
            mask = (mask_array == label_id)
            if not np.any(mask):
                continue
            volume = np.sum(mask) * voxel_volume
            mean_hu = np.mean(ct_array[mask])
            
            if organ in lession_volume_dict:
                row = [organ.replace('_', ' '), f"{volume:.2f}", f"{mean_hu:.1f}"]
            else:
                row = [organ.replace('_', ' '), f"{volume:.2f}", f"{mean_hu:.1f}"]
            draw_table_row(row)

        # y_position -= section_spacing

        # === Step 2: PDAC Staging ===
        # temp_pdf.setFont("Helvetica-Bold", 12)
        # temp_pdf.drawString(left_margin, y_position, "PDAC STAGING")
        # y_position -= line_height

        # try:
        #     pdac_info = get_pdac_staging(id)
        #     print(pdac_info, id)
        #     pdac_text = pdac_info.get("staging_description", "No staging data available.")
        # except Exception:
        #     pdac_text = "Error fetching PDAC staging information."

        # y_position = write_wrapped_text(left_margin, y_position, pdac_text, bold=False, font_size=10)
        # === Step 3: Key Images ===
        
        # include_liver = np.count_nonzero(mask_array == LABELS["liver"]) > 0
        # include_pancreas = lession_volume_dict.get("pancreas", {}).get("number", 0) > 0
        # include_kidney = np.count_nonzero(mask_array == LABELS["kidney_left"]) > 0 or np.count_nonzero(mask_array == LABELS["kidney_right"]) > 0
        # print(include_liver, include_pancreas, include_kidney)
        # if include_liver or include_pancreas or include_kidney:
        #     def check_and_reset_page(space_needed):
        #         nonlocal y_position
        #         if y_position - space_needed < 50:
        #             reset_page()

        #     temp_pdf.showPage()
        #     y_position = height - top_margin
        #     temp_pdf.setFont("Helvetica-Bold", 14)
        #     # temp_pdf.drawString(left_margin, y_position, "KEY IMAGES")
        #     y_position -= section_spacing

        #     organs = {
        #         "liver": include_liver,
        #         "pancreas": include_pancreas,
        #         "kidney_left": include_kidney,
        #         "kidney_right": include_kidney
        #     }
            # download_clean_folder(ct_path.replace("/inputs/", "/outputs/").rsplit("/", 1)[0])
            # for organ in organs:
            #     organ_data = lession_volume_dict.get(organ)
            #     if not organ_data or organ_data.get("number", 0) == 0:
            #         continue

            #     header = f"{organ.replace('_', ' ').upper()} TUMORS"
            #     check_and_reset_page(line_height)
            #     temp_pdf.setFont("Helvetica", 12)
            #     temp_pdf.drawString(left_margin, y_position, header)
            #     y_position -= line_height
            #     print(organ, organ_data)
            #     check_and_reset_page(220)
            #     overlay_path = f"/tmp/{organ}_overlay.png"
            #     print(ct_path, mask_path)
            #     organ_mask_path = mask_path.replace('combined_labels.nii.gz', 'segmentations/'+organ+'.nii.gz')
            #     print(organ_mask_path)
            #     if create_overlay_image(ct_path, organ_mask_path, overlay_path, color="red"):
            #         try:
            #             temp_pdf.drawImage(overlay_path, left_margin, y_position - 200, width=200, height=200)
            #         except:
            #             print(overlay_path)
            #     check_and_reset_page(220)
            #     zoom_path = f"/tmp/{organ}_zoomed.png"
            #     if zoom_into_labeled_area(ct_path, organ_mask_path, zoom_path, color="red"):
            #         temp_pdf.drawImage(zoom_path, left_margin + 250, y_position - 205, width=210, height=210)
            #     print('521')
            #     y_position -= 220

        temp_pdf.save()

        # Merge with template
        template_reader =  PdfReader(template_pdf)
        content_reader = PdfReader(temp_pdf_path)
        writer = PdfWriter()

        for page in content_reader.pages:
            template_page = template_reader.pages[0]
            merged_page = PageObject.create_blank_page(
                width=template_page.mediabox.width,
                height=template_page.mediabox.height
            )
            merged_page.merge_page(template_page)
            merged_page.merge_page(page)
            writer.add_page(merged_page)

        with open(output_pdf, "wb") as f:
            writer.write(f)

                
    except Exception as e:
        raise RuntimeError(f"Error generating PDF for {folder_name}: {e}")
    finally:
        if os.path.exists(temp_pdf_path):
            os.remove(temp_pdf_path)

# Helper Function to Process CT and Mask
def get_most_labeled_slice(ct_path, mask_path, output_png, contrast_min=-150, contrast_max=250):
    """
    Load CT and mask, ensure RAS orientation, find the most labeled slice, and generate an overlay image.
    """

    try:
        import SimpleITK as sitk
        import matplotlib
        matplotlib.use('Agg')  # ✅ 关键：不再尝试调用 GUI

        import matplotlib.pyplot as plt

        # Load the CT scan and mask
        ct_scan = sitk.ReadImage(ct_path)
        print('543',mask_path)
        mask = sitk.ReadImage(mask_path)
        print(mask_path)
        # Reorient to RAS
        ct_scan = sitk.DICOMOrient(ct_scan, 'RAS')
        mask = sitk.DICOMOrient(mask, 'RAS')

        # Convert to numpy arrays
        ct_array = sitk.GetArrayFromImage(ct_scan)
        mask_array = sitk.GetArrayFromImage(mask)

        # Check for shape mismatches
        if ct_array.shape != mask_array.shape:
            raise ValueError(f"Shape mismatch: CT shape {ct_array.shape}, Mask shape {mask_array.shape}")

        # Find the slice with the most labels
        slice_sums = np.sum(mask_array, axis=(1, 2))
        most_labeled_slice_index = np.argmax(slice_sums)

        # Get the CT and mask slices
        ct_slice = ct_array[most_labeled_slice_index]
        mask_slice = mask_array[most_labeled_slice_index]

        # Apply mirroring
        ct_slice = np.fliplr(ct_slice)
        mask_slice = np.fliplr(mask_slice)

        # Apply contrast adjustment
        ct_slice = np.clip(ct_slice, contrast_min, contrast_max)
        ct_slice = (ct_slice - contrast_min) / (contrast_max - contrast_min) * 255
        ct_slice = ct_slice.astype(np.uint8)

        # Overlay mask contours on CT slice
        plt.figure(figsize=(6, 6))
        plt.imshow(ct_slice, cmap='gray', origin='lower')
        plt.contour(mask_slice, colors='red', linewidths=1)  # Use red contours for the mask
        plt.axis('off')
        plt.savefig(output_png, bbox_inches="tight", pad_inches=0)
        plt.close()
        print('586')
        return True
    except:
        return False

def create_overlay_image(ct_path, mask_path, output_path, color="red"):
    """
    Generate overlay images for most labeled slices using the unified RAS orientation logic.
    """
    return get_most_labeled_slice(ct_path, mask_path, output_path)


# Helper Function to Zoom into Labeled Area
def zoom_into_labeled_area(ct_path, mask_path, output_path, color="red"):
    """
    Create a zoomed-in view of the largest labeled area with consistent RAS orientation.
    """
    import SimpleITK as sitk
    import matplotlib.pyplot as plt
    try:
        # Load the CT scan and mask
        ct_scan = sitk.ReadImage(ct_path)
        mask = sitk.ReadImage(mask_path)

        # Reorient to RAS
        ct_scan = sitk.DICOMOrient(ct_scan, 'RAS')
        mask = sitk.DICOMOrient(mask, 'RAS')

        # Convert to numpy arrays
        ct_array = sitk.GetArrayFromImage(ct_scan)
        mask_array = sitk.GetArrayFromImage(mask)

        # Check for shape mismatches
        if ct_array.shape != mask_array.shape:
            raise ValueError(f"Shape mismatch: CT shape {ct_array.shape}, Mask shape {mask_array.shape}")

        # Find the slice with the most labels
        slice_sums = np.sum(mask_array, axis=(1, 2))
        largest_slice_idx = np.argmax(slice_sums)
        if slice_sums[largest_slice_idx] == 0:
            raise ValueError("No labeled area found in the mask.")

        # Get the mask slice and calculate the bounding box
        mask_slice = mask_array[largest_slice_idx]
        coords = np.array(np.where(mask_slice))
        min_row, max_row = np.min(coords[0]), np.max(coords[0])
        min_col, max_col = np.min(coords[1]), np.max(coords[1])
        padding = 20
        min_row = max(min_row - padding, 0)
        max_row = min(max_row + padding, mask_slice.shape[0])
        min_col = max(min_col - padding, 0)
        max_col = min(max_col + padding, mask_slice.shape[1])

        # Extract the zoomed region
        zoomed_image = ct_array[largest_slice_idx][min_row:max_row, min_col:max_col]
        zoomed_mask = mask_array[largest_slice_idx][min_row:max_row, min_col:max_col]

        # Apply mirroring
        zoomed_image = np.fliplr(zoomed_image)
        zoomed_mask = np.fliplr(zoomed_mask)

        # Apply contrast adjustment to the zoomed CT slice
        zoomed_image = np.clip(zoomed_image, -150, 250)
        zoomed_image = (zoomed_image + 150) / 400 * 255
        zoomed_image = zoomed_image.astype(np.uint8)

        # Save the zoomed-in image with overlay
        plt.figure(figsize=(6, 6))
        plt.imshow(zoomed_image, cmap="gray", origin="lower")
        plt.contour(zoomed_mask, colors=color, linewidths=1)
        plt.axis("off")
        plt.savefig(output_path, bbox_inches="tight")
        plt.close()
        return True
    except Exception as e:
        return False

def get_pdac_staging(clabel_id):
    try:
        subfolder = "ImageTr" if int(clabel_id) < 9000 else "ImageTe"
        label_subfolder = "LabelTr" if int(clabel_id) < 9000 else "LabelTe"
        main_nifti_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(clabel_id)}/{Constants.MAIN_NIFTI_FILENAME}"
        combined_labels_path = f"{Constants.PANTS_PATH}/data/{label_subfolder}/{get_panTS_id(clabel_id)}/{Constants.COMBINED_LABELS_NIFTI_FILENAME}"
        
        nifti_processor = NiftiProcessor(main_nifti_path, combined_labels_path)
        staging_result = nifti_processor.calculate_pdac_sma_staging()

        return {"staging_description": staging_result}

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": f"PDAC staging failed: {str(e)}"}

import json
def download_clean_folder(root):
    """
    如果文件正好匹配4个目标名，则删除其中两个，并将combined_labels.nii.gz根据dataset.json分割为独立器官文件。
    """
    target_files = {
        "combined_labels.nii.gz",
        "dataset.json",
        "plans.json",
        "predict_from_raw_data_args.json"
    }

    actual_files = set(os.listdir(root))
    if actual_files == target_files:
        # 删除 plans.json 和 predict_from_raw_data_args.json
        for fname in ["plans.json", "predict_from_raw_data_args.json"]:
            fpath = os.path.join(root, fname)
            if os.path.exists(fpath):
                os.remove(fpath)
                print(f"🗑️ Removed during zip: {fpath}")

        # 读取 dataset.json
        dataset_json_path = os.path.join(root, "dataset.json")
        with open(dataset_json_path, 'r') as f:
            dataset_info = json.load(f)

        labels = dataset_info["labels"]  # 获取标签名与ID的映射

        # 读取 combined_labels.nii.gz
        combined_path = os.path.join(root, "combined_labels.nii.gz")
        combined_img = nib.load(combined_path)
        combined_data = combined_img.get_fdata()
        affine = combined_img.affine

        # 创建 segmentations 文件夹
        seg_folder = os.path.join(root, "segmentations")
        os.makedirs(seg_folder, exist_ok=True)

        # 为每个标签生成单独的 mask 文件
        for label_name, label_value in labels.items():
            mask = (combined_data == label_value).astype(np.uint8)
            label_img = nib.Nifti1Image(mask, affine)
            out_path = os.path.join(seg_folder, f"{label_name}.nii.gz")
            nib.save(label_img, out_path)
            print(f"✅ Saved: {out_path}")
        os.remove(dataset_json_path)
    else:
        print("ℹ️ Folder content does not match the expected file set. Skipping cleanup and split.")
        
async def store_files(combined_labels_id):
    subfolder = "LabelTr" if int(combined_labels_id) < 9000 else "LabelTe" 
    image_subfolder = "ImageTr" if int(combined_labels_id) < 9000 else "ImageTe"

    def download(url, path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        headers = {"User-Agent": "Mozilla/5.0"}
        res = requests.get(url, stream=True, headers=headers, allow_redirects=True)
        if res.status_code == 200:
            with open(path, "wb") as f:
                for chunk in res.iter_content(1024):
                    f.write(chunk)
            print(f"Saved: {path}")
        else:
            print(f"Failed: {url} ({res.status_code})")

    # main CT
    image_url = f"https://huggingface.co/datasets/BodyMaps/iPanTSMini/resolve/main/image_only/{get_panTS_id(combined_labels_id)}/ct.nii.gz"
    image_path = f"{Constants.PANTS_PATH}/data/{image_subfolder}/{get_panTS_id(combined_labels_id)}/ct.nii.gz"
    download(image_url, image_path)

    # labels
    for label in list(Constants.PREDEFINED_LABELS.values()):
        mask_url = f"https://huggingface.co/datasets/BodyMaps/iPanTSMini/resolve/main/mask_only/{get_panTS_id(combined_labels_id)}/segmentations/{label}.nii.gz"
        mask_path = f"{Constants.PANTS_PATH}/data/{subfolder}/{get_panTS_id(combined_labels_id)}/segmentations/{label}.nii.gz"
        download(mask_url, mask_path)
        
META_FILE = f"{Constants.PANTS_PATH}/data/metadata.xlsx"
# ---------------------------
# Helpers
# ---------------------------
def _arg(name: str, default=None):
    return request.args.get(name, default)

def to_int(x) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        return None

def _to_float(x) -> Optional[float]:
    try:
        return float(x)
    except Exception:
        return None

def _to01_query(x) -> Optional[int]:
    if x is None: return None
    s = str(x).strip().lower()
    if s in ("1","true","yes","y"): return 1
    if s in ("0","false","no","n"): return 0
    return None

def _collect_list_params(names: List[str]) -> List[str]:
    out: List[str] = []
    for n in names:
        if n in request.args:
            out += request.args.getlist(n)
    tmp: List[str] = []
    for s in out:
        if "," in s:
            tmp += [t.strip() for t in s.split(",") if t.strip()]
        else:
            tmp.append(s.strip())
    return [t for t in tmp if t]

def _nan2none(v):
    try:
        if v is None: return None
        if pd.isna(v): return None
    except Exception:
        pass
    return v

def clean_json_list(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def _clean(v):
        if isinstance(v, (np.integer,)):  return int(v)
        if isinstance(v, (np.floating,)): return float(v)
        if isinstance(v, (np.bool_,)):    return bool(v)
        return v
    return [{k: _clean(v) for k, v in d.items()} for d in items]

def _canon_letters_digits(s: str) -> str:
    # 把 "LightSpeed16" 變成 "LightSpeed 16"
    s2 = re.sub(r"([A-Za-z])(\d)", r"\1 \2", s)
    s2 = re.sub(r"(\d)([A-Za-z])", r"\1 \2", s2)
    return re.sub(r"\s+", " ", s2).strip()

def canon_model(s: str) -> str:
    if not s: return ""
    base = str(s).strip()
    # 標準化空白/底線/大小寫
    low = re.sub(r"[_\-]+", " ", base).strip().lower()
    low = _canon_letters_digits(low)
    # 套用別名表
    if low in Constants.MODEL_ALIASES:
        return Constants.MODEL_ALIASES[low]
    # 沒有在別名表時：維持「字母數字分隔 + 每字首大寫」的安全格式
    spaced = _canon_letters_digits(base)
    # 常見廠牌固定大寫
    spaced = re.sub(r"(?i)^somatom", "SOMATOM", spaced)
    spaced = re.sub(r"(?i)^iqon", "IQon", spaced)
    return spaced

# ---------------------------
# Load & normalize
# ---------------------------
def _norm_cols(df_raw: pd.DataFrame) -> pd.DataFrame:
    """標準化欄位，產出搜尋/排序需要的衍生欄位。"""
    df = df_raw.copy()

    # ---- Case ID ----
    case_cols = ["PanTS ID", "PanTS_ID", "case_id", "id", "case", "CaseID"]
    def _first_nonempty(row, cols):
        for c in cols:
            if c in row.index and pd.notna(row[c]) and str(row[c]).strip():
                return str(row[c]).strip(), c
        return "", None

    cases, mapping = [], []
    for _, r in df.iterrows():
        s, c = _first_nonempty(r, case_cols)
        cases.append(s); mapping.append({"case": c} if c else {})
    df["__case_str"] = cases
    df["_orig_cols"] = mapping

    # ---- Tumor -> __tumor01 ----
    def _canon(s: str) -> str: return re.sub(r"[^a-z]+", "", str(s).lower())
    tumor_names = [c for c in df.columns if "tumor" in _canon(c)] or []
    tcol = tumor_names[0] if tumor_names else None

    def _to01_v(v):
        if pd.isna(v): return np.nan
        s = str(v).strip().lower()
        if s in ("1","yes","y","true","t"): return 1
        if s in ("0","no","n","false","f"): return 0
        try:
            iv = int(float(s))
            return 1 if iv == 1 else (0 if iv == 0 else np.nan)
        except Exception:
            return np.nan

    df["__tumor01"] = (df[tcol].map(_to01_v) if tcol else pd.Series([np.nan]*len(df), index=df.index))
    if tcol:
        df["_orig_cols"] = [{**(df["_orig_cols"].iat[i] or {}), "tumor": tcol} for i in range(len(df))]

    # ---- Sex -> __sex ----
    df["__sex"] = df.get("sex", pd.Series([""]*len(df))).astype(str).str.strip().str.upper()
    df["__sex"] = df["__sex"].where(df["__sex"].isin(["F","M"]), "")

    # ---- Generic column finder ----
    def _find_col(prefer, keyword_sets=None):
        for c in prefer:
            if c in df.columns: return c
        if keyword_sets:
            canon_map = {c: re.sub(r"[^a-z0-9]+", "", str(c).lower()) for c in df.columns}
            for c, cs in canon_map.items():
                for ks in keyword_sets:
                    if all(k in cs for k in ks): return c
        return None

    # ---- CT phase -> __ct / __ct_lc ----
    ct_col = _find_col(
        prefer=["ct phase","CT phase","ct_phase","CT_phase","ct"],
        keyword_sets=[["ct","phase"],["phase"]],
    )
    if ct_col:
        df["__ct"] = df[ct_col].astype(str).str.strip()
        df["__ct_lc"] = df["__ct"].str.lower()
        df["_orig_cols"] = [{**(df["_orig_cols"].iat[i] or {}), "ct_phase": ct_col} for i in range(len(df))]
    else:
        df["__ct"], df["__ct_lc"] = "", ""

    # ---- Manufacturer -> __mfr / __mfr_lc ----
    mfr_col = _find_col(
        prefer=["manufacturer","Manufacturer","mfr","MFR","vendor","Vendor","manufacturer name","Manufacturer Name"],
        keyword_sets=[["manufactur"],["vendor"],["brand"],["maker"]],
    )
    if mfr_col:
        df["__mfr"] = df[mfr_col].astype(str).str.strip()
        df["__mfr_lc"] = df["__mfr"].str.lower()
        df["_orig_cols"] = [{**(df["_orig_cols"].iat[i] or {}), "manufacturer": mfr_col} for i in range(len(df))]
    else:
        df["__mfr"], df["__mfr_lc"] = "", ""

        # ---- Manufacturer model -> model / __model_lc ----
    model_col = _find_col(
        prefer=["manufacturer model", "Manufacturer model", "model", "Model"],
        keyword_sets=[["model"]],
    )
    if model_col:
        # 保留原始字串以便追蹤
        df["model_raw"] = df[model_col].astype(str).str.strip()
        # 規則化為標準型號（大小寫、空白、數字黏在一起等）
        df["model"] = df["model_raw"].map(canon_model)
        df["__model_lc"] = df["model"].str.lower()
        df["_orig_cols"] = [
            {**(df["_orig_cols"].iat[i] or {}), "model": model_col}
            for i in range(len(df))
        ]
    else:
        # 以免前端讀不到欄位
        df["model_raw"] = ""
        df["model"] = ""
        df["__model_lc"] = ""

    # ---- Year -> __year_int ----
    year_col = _find_col(prefer=["study year", "Study year", "study_year", "year", "Year"],
                         keyword_sets=[["year"]])
    df["__year_int"] = (
        pd.to_numeric(df[year_col], errors="coerce")
        if year_col else pd.Series([np.nan] * len(df), index=df.index)
    )
    if year_col:
        df["_orig_cols"] = [
            {**(df["_orig_cols"].iat[i] or {}), "year": year_col}
            for i in range(len(df))
        ]

    # ---- Age -> __age ----
    age_col = _find_col(prefer=["age", "Age"], keyword_sets=[["age"]])
    df["__age"] = (
        pd.to_numeric(df[age_col], errors="coerce")
        if age_col else pd.Series([np.nan] * len(df), index=df.index)
    )
    if age_col:
        df["_orig_cols"] = [
            {**(df["_orig_cols"].iat[i] or {}), "age": age_col}
            for i in range(len(df))
        ]

    # ---- Study type -> study_type / __st_lc ----
    st_col = _find_col(
        prefer=["study type", "Study type", "study_type", "Study_type"],
        keyword_sets=[["study", "type"]],
    )
    if st_col:
        df["study_type"] = df[st_col].astype(str)
        df["__st_lc"] = df["study_type"].astype(str).str.strip().str.lower()
        df["_orig_cols"] = [
            {**(df["_orig_cols"].iat[i] or {}), "study_type": st_col}
            for i in range(len(df))
        ]
    else:
        df["study_type"] = ""
        df["__st_lc"] = ""

    # ---- Site nationality -> site_nationality / __sn_lc ----
    sn_col = _find_col(
        prefer=[
            "site nationality", "Site nationality", "site_nationality", "Site_nationality",
            "nationality", "Nationality", "site country", "Site country", "country", "Country"
        ],
        keyword_sets=[["site", "national"], ["nationality"], ["site", "country"], ["country"]],
    )
    if sn_col:
        df["site_nationality"] = df[sn_col].astype(str)
        df["__sn_lc"] = df["site_nationality"].astype(str).str.strip().str.lower()
        df["_orig_cols"] = [
            {**(df["_orig_cols"].iat[i] or {}), "site_nationality": sn_col}
            for i in range(len(df))
        ]
    else:
        df["site_nationality"] = ""
        df["__sn_lc"] = ""

    return df


def _safe_float(x) -> Optional[float]:
    try:
        if x is None: return None
        if isinstance(x, float) and np.isnan(x): return None
        if isinstance(x, str):
            s = x.strip().replace(",", " ")
            if not s: return None
            return float(s)
        return float(x)
    except Exception:
        return None

def _take_first_str(row, cols: List[str]) -> str:
    for c in cols:
        if c in row and pd.notna(row[c]) and str(row[c]).strip():
            return str(row[c]).strip()
    return ""

def _case_key(row) -> int:
    s = _take_first_str(row, ["PanTS ID","PanTS_ID","case_id","id","__case_str"])
    if not s: return 0
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else 0

def _parse_3tuple_from_row(row, name_candidates: List[str]) -> List[Optional[float]]:
    # 3 個獨立欄
    for base in name_candidates:
        cx, cy, cz = f"{base}_x", f"{base}_y", f"{base}_z"
        if cx in row and cy in row and cz in row:
            xs = [_safe_float(row[c]) for c in (cx, cy, cz)]
            if all(v is not None for v in xs):
                return xs
    # 單欄字串
    seps = [",", "x", " ", "×", "X", ";", "|"]
    str_cols = []
    for base in name_candidates:
        str_cols += [base, f"{base}_str", base.replace(" ", "_")]
    for c in str_cols:
        if c in row and pd.notna(row[c]):
            s = str(row[c]).strip()
            if not s: continue
            s2 = re.sub(r"[\[\]\(\)\{\}]", " ", s)
            for sep in seps:
                s2 = s2.replace(sep, " ")
            parts = [p for p in s2.split() if p]
            vals = [_safe_float(p) for p in parts[:3]]
            if len(vals) == 3 and all(v is not None for v in vals):
                return vals
    return [None, None, None]

def _spacing_sum(row) -> Optional[float]:
    vals = _parse_3tuple_from_row(row, ["spacing","voxel_spacing","voxel_size","pixel_spacing"])
    if any(v is None for v in vals): return None
    return float(vals[0] + vals[1] + vals[2])

def _shape_sum(row) -> Optional[float]:
    vals = _parse_3tuple_from_row(row, ["shape","dim","size","image_shape","resolution"])
    if any(v is None for v in vals): return None
    return float(vals[0] + vals[1] + vals[2])

def ensure_sort_cols(df: pd.DataFrame) -> pd.DataFrame:
    if "__case_sortkey" not in df.columns:
        df["__case_sortkey"] = df.apply(_case_key, axis=1)
    if "__spacing_sum" not in df.columns:
        df["__spacing_sum"] = df.apply(_spacing_sum, axis=1)
    if "__shape_sum" not in df.columns:
        df["__shape_sum"] = df.apply(_shape_sum, axis=1)

    # 完整度：Browse 與 top 排序會用到
    need_cols = ["__spacing_sum", "__shape_sum", "__sex", "__age"]
    complete = pd.Series(True, index=df.index)
    for c in need_cols:
        if c not in df.columns:
            complete &= False
        elif c == "__sex":
            complete &= (df[c].astype(str).str.strip() != "")
        else:
            complete &= df[c].notna()
    df["__complete"] = complete
    return df

# load meta
if not os.path.exists(META_FILE):
    raise FileNotFoundError(f"metadata not found: {META_FILE}")
DF_RAW = pd.read_excel(META_FILE)
DF = _norm_cols(DF_RAW)

def apply_filters(base: pd.DataFrame, exclude: Optional[Set[str]] = None) -> pd.DataFrame:
    exclude = exclude or set()
    df = base

    # --- Case ID / keyword（精準匹配） ---
    q = (_arg("q") or _arg("caseid") or "").strip()
    if q and "caseid" not in exclude and "__case_str" in df.columns:
        s = df["__case_str"].astype(str)
        if q.isdigit():
            # 把每列所有數字 token 抓出來，做數值等號；77 不會吃 177/077（前導 0 忽略）
            qq = int(q)
            nums = s.str.findall(r"\d+")
            mask_num = nums.apply(lambda xs: any(int(x) == qq for x in xs))
            # 備援：允許 "Case 77"（不必留可刪）
            patt = rf"(?i)\b(?:case\s*)?{re.escape(q)}\b"
            mask_regex = s.str.contains(patt, na=False, regex=True)
            df = df[mask_num | mask_regex]
        else:
            # 一般文字搜尋（忽略大小寫；避免把查詢當正則）
            df = df[s.str.contains(re.escape(q), na=False, case=False, regex=False)]

    # --- Tumor ---
    tv = _to01_query(_arg("tumor"))
    tnull = _to01_query(_arg("tumor_is_null"))
    if (_arg("tumor", "").strip().lower() == "unknown"):
        tnull, tv = 1, None
    if "__tumor01" in df.columns and "tumor" not in exclude:
        if tnull in (0, 1) and "tumor_is_null" not in exclude:
            df = df[df["__tumor01"].isna()] if tnull == 1 else df[df["__tumor01"].notna()]
        elif tv in (0, 1):
            df = df[df["__tumor01"] == tv]

    # --- Sex（多選 + Unknown）---
    sv_list = _collect_list_params(["sex", "sex[]"])
    snull = _to01_query(_arg("sex_is_null"))
    if not sv_list:
        sv = (_arg("sex", "") or "").strip().upper()
        if sv:
            sv_list = [sv]
    sv_norm = []
    for s_ in sv_list:
        s2 = (s_ or "").strip().upper()
        if s2 in ("M", "F"):
            sv_norm.append(s2)
        elif s2 in ("U", "UNKNOWN"):
            sv_norm.append("UNKNOWN")
    if "__sex" in df.columns and "sex" not in exclude and (sv_norm or snull in (0, 1)):
        ser = df["__sex"].fillna("").str.strip().str.upper()
        take = pd.Series(False, index=df.index)
        vals = [s for s in sv_norm if s in ("F", "M")]
        if vals:
            take |= ser.isin(vals)
        if ("UNKNOWN" in sv_norm) or (snull == 1):
            take |= (ser == "")
        df = df[take]

    # --- Age：支援 age_bin[]（含 90+ / UNKNOWN），否則回退 age_from/age_to ---
    bins = _collect_list_params(["age_bin", "age_bin[]"])
    age_null = _to01_query(_arg("age_is_null"))
    if "__age" in df.columns and bins:
        age_series = pd.to_numeric(df["__age"], errors="coerce")
        mask = pd.Series(False, index=df.index)
        for b in bins:
            s = (b or "").strip()
            m_plus = re.match(r"^\s*(\d+)\s*\+\s*$", s)
            if m_plus:
                lo = int(m_plus.group(1))
                mask |= (age_series >= lo)
                continue
            m_rng = re.match(r"^\s*(\d+)\s*[-–—]\s*(\d+)\s*$", s)
            if m_rng:
                lo, hi = int(m_rng.group(1)), int(m_rng.group(2))
                mask |= age_series.between(lo, hi, inclusive="both")
        if (age_null == 1) or any((t or "").strip().upper() == "UNKNOWN" for t in bins):
            mask |= age_series.isna() | (df["__age"].astype(str).str.strip().str.upper() == "UNKNOWN")
        df = df[mask]
    elif "__age" in df.columns:
        af = _to_float(_arg("age_from")); at = _to_float(_arg("age_to"))
        age_series = pd.to_numeric(df["__age"], errors="coerce")
        if "age_from" not in exclude and af is not None:
            df = df[age_series >= af]
        if "age_to" not in exclude and at is not None:
            df = df[age_series <= at]

    # --- CT phase ---
    ct = (_arg("ct_phase", "") or "").strip().lower()
    ct_list = _collect_list_params(["ct_phase", "ct_phase[]"])
    if ct == "unknown" or any((s or "").lower() == "unknown" for s in ct_list):
        if "__ct" in df.columns:
            s_ct = df["__ct"].astype(str).str.strip().str.lower()
            tokens_null_ct = {'', 'unknown', 'nan', 'n/a', 'na', 'none', '(blank)', '(null)'}
            df = df[df["__ct"].isna() | s_ct.isin(tokens_null_ct)]
    elif (ct or ct_list) and "__ct_lc" in df.columns:
        parts = []
        if ct:
            parts += [p.strip() for p in re.split(r"[;,/]+", ct) if p.strip()]
        parts += [p.strip().lower() for p in ct_list if p.strip()]
        patt = "|".join(re.escape(p) for p in parts)
        df = df[df["__ct_lc"].str.contains(patt, na=False)]

    # --- Manufacturer ---
    m_list = _collect_list_params(["manufacturer", "manufacturer[]", "mfr"])
    m_raw = (_arg("manufacturer", "") or "").strip()
    if m_raw and not m_list:
        m_list = [p.strip() for p in m_raw.split(",") if p.strip()]
    if m_list and "__mfr_lc" in df.columns:
        m_lc = [s.lower() for s in m_list]
        df = df[df["__mfr_lc"].isin(m_lc)]

    # --- Model（canonical；可 fuzzy）---
    model_list = _collect_list_params(["model", "model[]", "manufacturer_model"])
    model_raw = (_arg("model", "") or "").strip()
    if model_raw and not model_list:
        model_list = [p.strip() for p in re.split(r"[;,/|]+", model_raw) if p.strip()]
    if model_list and "__model_lc" in df.columns and "model" not in exclude:
        wants = [canon_model(p).lower() for p in model_list if p]
        wants = [w for w in wants if w]
        fuzzy = str(_arg("model_fuzzy", "0")).lower() in ("1", "true", "yes")
        if fuzzy:
            patt = "|".join(re.escape(w) for w in wants)
            df = df[df["__model_lc"].str.contains(patt, na=False)]
        else:
            df = df[df["__model_lc"].isin(set(wants))]

    # --- Study type ---
    st_list = _collect_list_params(["study_type", "study_type[]"])
    st_raw = (_arg("study_type", "") or "").strip()
    if st_raw and not st_list:
        st_list = [p.strip() for p in re.split(r"[;,/|]+", st_raw) if p.strip()]
    if st_list and "__st_lc" in df.columns and "study_type" not in exclude:
        parts = [p.lower() for p in st_list]
        patt = "|".join(re.escape(p) for p in parts)
        df = df[df["__st_lc"].str.contains(patt, na=False)]

    # --- Site nationality ---
    nat_list = _collect_list_params(["site_nat", "site_nat[]", "site_nationality", "site_nationality[]"])
    nat_raw = (_arg("site_nationality", "") or _arg("site_nat", "") or "").strip()
    if nat_raw and not nat_list:
        nat_list = [p.strip() for p in re.split(r"[;,/|]+", nat_raw) if p.strip()]
    if nat_list and "__sn_lc" in df.columns and "site_nationality" not in exclude:
        parts = [p.lower() for p in nat_list]
        patt = "|".join(re.escape(p) for p in parts)
        df = df[df["__sn_lc"].str.contains(patt, na=False)]

    # --- Year（新增）---
    # 支援 year / year[]（多選精確）、year_from / year_to（範圍）與 year_is_null（Unknown）
    if "year" not in exclude:
        _year_cols_pref = ["__year_int", "study_year", "Study year", "study year", "Year", "year"]
        _found_cols = [c for c in _year_cols_pref if c in df.columns]
        if _found_cols:
            yser = pd.to_numeric(df[_found_cols[0]], errors="coerce")

            # 1) 多選年份
            year_list = _collect_list_params(["year", "year[]"])
            year_raw = (_arg("year", "") or "").strip()
            if year_raw and not year_list:
                year_list = [p.strip() for p in re.split(r"[;,/|]+", year_raw) if p.strip()]

            # 2) 範圍
            y_from = to_int(_arg("year_from"))
            y_to   = to_int(_arg("year_to"))

            # 3) Unknown / Null
            y_is_null = _to01_query(_arg("year_is_null"))
            _unk_tokens = {"unknown", "nan", "none", "n/a", "na", "(blank)", "(null)"}
            wants_unknown = (y_is_null == 1) or any(
                (s or "").strip().lower() in _unk_tokens for s in year_list
            )

            mask = pd.Series(True, index=df.index)

            # 多選精確年份
            exact_years = []
            for s in year_list:
                try:
                    exact_years.append(int(s))
                except Exception:
                    pass
            if exact_years:
                mask &= yser.isin(set(exact_years))

            # 範圍條件
            if y_from is not None:
                mask &= (yser >= y_from)
            if y_to is not None:
                mask &= (yser <= y_to)

            # Unknown 合併進來
            if wants_unknown:
                mask = mask | yser.isna()

            df = df[mask]


    return df

def row_to_item(row: pd.Series) -> Dict[str, Any]:
    cols = row.get("_orig_cols")
    cols = cols if isinstance(cols, dict) else {}

    def pick(k, fallback=None):
        col = cols.get(k)
        if col and col in row.index:
            return row[col]
        return fallback

    return {
        "PanTS ID": _nan2none(pick("case") or row.get("__case_str")),
        "case_id":  _nan2none(pick("case") or row.get("__case_str")),
        "tumor":    (int(row.get("__tumor01")) if pd.notna(row.get("__tumor01")) else None),
        "sex":      _nan2none(row.get("__sex")),
        "age":      _nan2none(row.get("__age")),
        "ct phase": _nan2none(pick("ct_phase") or row.get("__ct")),
        "manufacturer": _nan2none(pick("manufacturer") or row.get("__mfr")),
        "manufacturer model": _nan2none(pick("model") or row.get("model")),
        "study year": _nan2none(row.get("__year_int")),
        "study type": _nan2none(pick("study_type") or row.get("study_type")),
        "site nationality": _nan2none(pick("site_nationality") or row.get("site_nationality")),
        # 排序輔助輸出
        "spacing_sum": _nan2none(row.get("__spacing_sum")),
        "shape_sum":   _nan2none(row.get("__shape_sum")),
        "complete":    bool(row.get("__complete")) if "__complete" in row else None,
    }
